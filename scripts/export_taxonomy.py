"""
Export GAM taxonomy data from XLSX → JSON seed files for Laravel.
Run: python scripts/export_taxonomy.py
"""
import openpyxl
import json
import os

XLSX = r'C:\Users\hp\OneDrive - FAST National University\Cyber Security\Sem-4\Internship\GAM\GAM - ALL SHEETS FINAL VER_UPDATED_RECALC.xlsx'
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'database', 'data')
os.makedirs(OUT, exist_ok=True)

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ─── 1. Primary Groups ──────────────────────────────────────────────────────
ws = wb['Primary Group']
groups = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    code, label, desc, steward, notes = (row + (None,)*5)[:5]
    if not code:
        continue
    groups.append({
        'term_type': 'primary_group',
        'group_code': str(code).strip(),
        'term_code': str(code).strip(),
        'term_label': str(label).strip() if label else str(code).strip(),
        'description': str(desc).strip() if desc else '',
        'parent_code': None,
        'facet': None,
    })

# ─── 2. Doc Groups ──────────────────────────────────────────────────────────
ws = wb['Docs - Primary Groups']
doc_groups = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    code, label, desc, types, access, ai_proc = (row + (None,)*6)[:6]
    if not code:
        continue
    doc_groups.append({
        'term_type': 'doc_group',
        'group_code': str(code).strip(),
        'term_code': str(code).strip(),
        'term_label': str(label).strip() if label else str(code).strip(),
        'description': str(desc).strip() if desc else '',
        'parent_code': None,
        'facet': None,
        'extra': {
            'file_types': str(types).strip() if types else '',
            'access_level': str(access).strip() if access else '',
            'ai_processing': str(ai_proc).strip() if ai_proc else '',
        }
    })

# ─── 3. Project Tag Packs ───────────────────────────────────────────────────
projects = []
for sheet_name, parent_group in [('Food Group Tag Pack', 'FOOD'), ('Media Group Tag Pack', 'MEDIA')]:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        code, label, parent, notes = (row + (None,)*4)[:4]
        if not code or str(code).startswith('NOTE'):
            continue
        projects.append({
            'term_type': 'project',
            'group_code': parent_group,
            'term_code': str(code).strip(),
            'term_label': str(label).strip() if label else '',
            'description': str(notes).strip() if notes else '',
            'parent_code': parent_group if str(parent).strip() == '(root)' else str(parent).strip() if parent else None,
            'facet': None,
        })

# ─── 4. Visual Keyword Taxonomies ───────────────────────────────────────────
keyword_sheets = {
    'Food Keywords Taxonomy': 'FOOD',
    'Media Keywords Taxonomy': 'MEDIA',
    'Gen Business Keywords Taxonomy': 'GENBUS',
    'Geo Keywords Taxonomy': 'GEO',
    'Nature Keywords Taxonomy': 'NATURE',
    'Lifestyle Keywords Taxonomy': 'LIFE',
    'Specialty Keywords Taxonomy': 'SPEC',
}

keywords = []
for sheet_name, group_code in keyword_sheets.items():
    ws = wb[sheet_name]
    current_category = None
    current_mid = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        mid, sid, label, notes = (row + (None,)*4)[:4]
        
        if not label:
            continue
        label_str = str(label).strip()
        if not label_str:
            continue
        
        mid_str = str(mid).strip() if mid else ''
        sid_str = str(sid).strip() if sid else ''
        
        # Top-level category (has MID letter like A, B, C...)
        if mid_str and mid_str.isalpha() and len(mid_str) == 1:
            current_category = label_str
            current_mid = mid_str
            keywords.append({
                'term_type': 'keyword_category',
                'group_code': group_code,
                'term_code': f'{group_code}_{mid_str}',
                'term_label': label_str,
                'description': str(notes).strip() if notes else '',
                'parent_code': group_code,
                'facet': 'category',
            })
            continue
        
        # Sub-category (has SID number)
        if sid_str and sid_str.isdigit():
            keywords.append({
                'term_type': 'keyword',
                'group_code': group_code,
                'term_code': f'{group_code}_{current_mid}_{sid_str}' if current_mid else f'{group_code}_{sid_str}',
                'term_label': label_str,
                'description': str(notes).strip() if notes else '',
                'parent_code': f'{group_code}_{current_mid}' if current_mid else group_code,
                'facet': current_category,
            })
            continue
        
        # Leaf term (no MID/SID — nested under last SID)
        if label_str and not mid_str and not sid_str:
            keywords.append({
                'term_type': 'keyword',
                'group_code': group_code,
                'term_code': None,  # auto-generate
                'term_label': label_str,
                'description': str(notes).strip() if notes else '',
                'parent_code': None,  # last parent
                'facet': current_category,
            })

# ─── 5. Doc Keywords Taxonomy ───────────────────────────────────────────────
ws = wb['Docs Keywords Taxonomy']
doc_keywords = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    doc_group, doc_label, facet, term_code, term, synonyms, notes = (row + (None,)*7)[:7]
    if not term:
        continue
    doc_keywords.append({
        'term_type': 'doc_keyword',
        'group_code': str(doc_group).strip() if doc_group else '',
        'term_code': str(term_code).strip() if term_code else '',
        'term_label': str(term).strip(),
        'description': str(notes).strip() if notes else '',
        'parent_code': str(doc_group).strip() if doc_group else None,
        'facet': str(facet).strip() if facet else '',
        'extra': {
            'synonyms': str(synonyms).strip() if synonyms else '',
        }
    })

# ─── 6. VIZ Keywords ────────────────────────────────────────────────────────
ws = wb['VIZ Keywords Taxonomy']
viz_keywords = []
current_facet = None
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    facet, term, code, definition, synonyms, donot, example = (row + (None,)*7)[:7]
    if facet and str(facet).strip():
        current_facet = str(facet).strip()
    if not term:
        continue
    viz_keywords.append({
        'term_type': 'viz_keyword',
        'group_code': 'VIZ',
        'term_code': str(code).strip() if code else '',
        'term_label': str(term).strip(),
        'description': str(definition).strip() if definition else '',
        'parent_code': 'VIZ',
        'facet': current_facet,
        'extra': {
            'synonyms': str(synonyms).strip() if synonyms else '',
            'do_not_use': str(donot).strip() if donot else '',
            'example': str(example).strip() if example else '',
        }
    })

# ─── 7. Geo Crosswalk ───────────────────────────────────────────────────────
ws = wb['Geo Crosswalk']
geo = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    state, usps, region1, region2, census, notes = (row + (None,)*6)[:6]
    if not state:
        continue
    geo.append({
        'term_type': 'geo_state',
        'group_code': 'GEO',
        'term_code': str(usps).strip() if usps else '',
        'term_label': str(state).strip(),
        'description': '',
        'parent_code': str(region1).strip() if region1 else None,
        'facet': 'State',
        'extra': {
            'region_primary': str(region1).strip() if region1 else '',
            'region_secondary': str(region2).strip() if region2 else '',
            'census_region': str(census).strip() if census else '',
        }
    })

# ─── 8. Synonyms ────────────────────────────────────────────────────────────
synonym_sheets = {
    'Synonyms - Normalization': None,
    'Synonyms - Geo Markets (Seed)': 'GEO',
    'Synonyms - Food (Seed)': 'FOOD',
    'Synonyms - Media (Seed)': 'MEDIA',
    'Synonyms - Docs Acronyms (Seed)': 'DOC',
}

synonyms = []
for sheet_name, default_group in synonym_sheets.items():
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if sheet_name == 'Synonyms - Normalization':
            domain, category, preferred, variant, normalize_to, notes = (row + (None,)*6)[:6]
            if not variant or not preferred:
                continue
            # Determine group hint from domain/category
            group = None
            cat_str = str(category).strip() if category else ''
            if 'Primary Group' in cat_str:
                group = None  # cross-cutting
            elif 'Food' in cat_str:
                group = 'FOOD'
            elif 'Media' in cat_str:
                group = 'MEDIA'
            elif 'Geo' in cat_str:
                group = 'GEO'
            elif 'Visual' in cat_str:
                group = None  # cross-cutting attribute
            synonyms.append({
                'raw_term': str(variant).strip(),
                'canonical_term': str(normalize_to or preferred).strip(),
                'group_hint': group,
                'category': str(category).strip() if category else '',
            })
        else:
            # Seed sheets: typically domain, category, preferred, variant, normalize_to, notes
            cols = (row + (None,)*6)[:6]
            domain, category, preferred, variant, normalize_to, notes = cols
            if not variant or not preferred:
                continue
            synonyms.append({
                'raw_term': str(variant).strip(),
                'canonical_term': str(normalize_to or preferred).strip(),
                'group_hint': default_group,
                'category': str(category).strip() if category else '',
            })

# ─── Write JSON files ────────────────────────────────────────────────────────
all_terms = groups + doc_groups + projects + keywords + doc_keywords + viz_keywords + geo

with open(os.path.join(OUT, 'taxonomy_terms.json'), 'w', encoding='utf-8') as f:
    json.dump(all_terms, f, indent=2, ensure_ascii=False)

with open(os.path.join(OUT, 'taxonomy_synonyms.json'), 'w', encoding='utf-8') as f:
    json.dump(synonyms, f, indent=2, ensure_ascii=False)

# Summary
print(f'Exported to {OUT}:')
print(f'  taxonomy_terms.json:    {len(all_terms)} terms')
print(f'    - Primary groups:     {len(groups)}')
print(f'    - Doc groups:         {len(doc_groups)}')
print(f'    - Projects:           {len(projects)}')
print(f'    - Visual keywords:    {len(keywords)}')
print(f'    - Doc keywords:       {len(doc_keywords)}')
print(f'    - VIZ keywords:       {len(viz_keywords)}')
print(f'    - Geo states:         {len(geo)}')
print(f'  taxonomy_synonyms.json: {len(synonyms)} synonym mappings')
print(f'  TOTAL: {len(all_terms) + len(synonyms)} entries')
